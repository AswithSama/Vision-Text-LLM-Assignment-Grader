from pathlib import Path

from openpyxl import Workbook, load_workbook


RESULTS_FILE = Path("data/results/grading_results.xlsx")


HEADERS = [
    "Student ID",
    "Question ID",
    "Grade",
    "Maximum Points",
    "Grading Method",
    "Student Answer",
    "Reason",
]


def save_results_to_excel(
    student_name,
    grading_results,
    extracted_answers,
):
    RESULTS_FILE.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    if RESULTS_FILE.exists():
        workbook = load_workbook(RESULTS_FILE)
        worksheet = workbook.active
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Grades"
        worksheet.append(HEADERS)

    for question_id, grading_result in grading_results.items():
        worksheet.append(
            [
                student_name,
                question_id,
                grading_result["grade"],
                grading_result["max_points"],
                grading_result["grading_method"],
                extracted_answers[question_id]["combined_text"],
                grading_result["reason"],
            ]
        )

    workbook.save(RESULTS_FILE)

    print(
        f"Grading results stored in {RESULTS_FILE}"
    )